import numpy as np
import time
from pathlib import Path
from openpyxl import Workbook, load_workbook
import os
# custom imports
from file_handling import get_global_files, list_gerbers_with_weights, wait_for_folder_complete, clear_folder
from calculations.transformation import *
from loading.gerber_conversions import gerber_to_png_gerbv
from loading.img2array import bitmap_to_array
from calculations.multi_layer import multiple_layers_weighted
from calculations.layer_calcs import met_ave
from plotting.comparing import plot_pointclouds_and_heatmaps
from calculations.padding import fill_border
from calculations.comparison import align_and_compare
from pathlib import Path

project_folder = Path.cwd()

top_data_loc = str(project_folder / "Assets" / "AkroFiles" / "TopDatFiles")  # The folder that holds the top Akro data
bot_data_loc = str(project_folder / "Assets" / "AkroFiles" / "BottomDatFiles")  # The folder that holds the bottom Akro data

temp_tiff_folder = str(project_folder / "Assets" / "temp_tiff")
temp_tiff_folder_path = Path(temp_tiff_folder)


Q1_folder = str(project_folder / "Assets" / "gerbers" / "Cu_Balancing_Gerber" / "Q1")
Q3_folder = str(project_folder / "Assets" / "gerbers" / "Cu_Balancing_Gerber" / "Q3")
Q2_folder = str(project_folder / "Assets" / "gerbers" / "Cu_Balancing_Gerber" / "Q2")
Q4_folder = str(project_folder / "Assets" / "gerbers" / "Cu_Balancing_Gerber" / "Q4")
dpi_results = [100, 150, 200, 400, 450, 500, 550, 600, 650, 700]
fills = ['max_percent', 'mean_percent', 'biharmonic','idw','nearest',   'local_mean']
radii = [75, 200, 400, 600]
# Create the excel file
excel_output_path = str(project_folder / "Assets" / "Output" / "results.xlsx")
if not os.path.exists(excel_output_path):
    wb = Workbook()
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ws.append([
        "Identifier", "quartile", "edge_fill", "dpi", "radius", "mat_sup_id", "scale",
        "n_points", "rmse", "mae", "bias", "pearson_r", "spearman_r", "r2",
        "svd_var_explained", "slope", "intercept", "lambda_ratio", "notes"
    ])
    wb.save(excel_output_path)


Quartile_loc = ""

if __name__ == '__main__':

    # Load the Workbook
    wb = load_workbook(excel_output_path)
    ws = wb.active

    # Read in all global DAT forms
    top_data_files = get_global_files(top_data_loc)
    bot_data_files = get_global_files(bot_data_loc)
    # Gerber Files
    Q1_Gerber_files = list_gerbers_with_weights(Q1_folder)
    Q3_Gerber_files = list_gerbers_with_weights(Q3_folder)
    Q2_Gerber_files = list_gerbers_with_weights(Q2_folder)
    Q4_Gerber_files = list_gerbers_with_weights(Q4_folder)
    # Cycle through the dpi

    for dpi in dpi_results:
        for edge_fill in fills:

            for radius in radii:
                arrays = {}
                # Cycle through the Top Global Data Files
                for top_global_path in top_data_files:
                    # Material and Supplier
                    tmp_id = '-'.join(top_global_path.split('\\')[-1].split('-')[0:3])
                    mat_sup_id = f"{tmp_id}_EF{edge_fill}_DPI{dpi}_R{radius}"
                    mat_sup_folder = '-'.join(top_global_path.split('\\')[-1].split('-')[0:2])
                    # Clear the temporary tiff folder
                    # print("Clearing Temporary Tiff Folder")
                    clear_folder(temp_tiff_folder)
                    time.sleep(.1)
                    # Create list of layer names to be blended
                    layer_names_for_blend = []
                    layer_weights_for_blend = {}
                    recalculate_array = False  # If we have the array already stored in memory, no need to recalculate anything


                    #Wait for calculations
                    wait_for_calcs = False
                    if "Q1" in top_global_path:
                        Quartile_loc = "Q1"
                        gerber_files = Q1_Gerber_files
                    elif "Q2" in top_global_path:
                        Quartile_loc = "Q2"
                        gerber_files = Q2_Gerber_files
                    elif "Q3" in top_global_path:
                        Quartile_loc = "Q3"
                        gerber_files = Q3_Gerber_files
                    elif "Q4" in top_global_path:
                        Quartile_loc = "Q4"
                        gerber_files = Q4_Gerber_files
                    else:
                        gerber_files = []

                    plot_save_folder = str(project_folder / "Assets" / "Output" / f"{mat_sup_folder}" / f"{Quartile_loc}")
                    plot_save_name = f'{Quartile_loc}_{mat_sup_id}'
                    if Path('\\'.join([plot_save_folder, plot_save_name])+".html").exists():
                        print(f"Completed Prior: {Quartile_loc}_{mat_sup_id}")
                        continue
                    # Read the Akro Arrays and interpolate the nan values so we dont have 9999 or np.nan
                    try:
                        dat_file_orig = np.loadtxt(top_global_path, delimiter="\t")
                        dat_file_filled = np.where(dat_file_orig == 9999.0, np.nan, dat_file_orig)
                        dat_file_9999_filled = fill_nans_nd(dat_file_filled, 'iterative')
                    except Exception as error:
                        print("Error Reading DAT: ", error)
                        continue

                    for gerber_path in gerber_files:
                        name = gerber_path[0].replace('/', '\\').split("\\")[-1].split(".gbr")[0]
                        layer_names_for_blend.append(name)
                        layer_weights_for_blend[name] = gerber_path[1]
                        if name in arrays:
                            continue
                        gerber_to_png_gerbv(gerb_file=gerber_path[0], save_folder_temp=temp_tiff_folder, save_name=name, dpi=dpi)  # Convert the gerbers to arrays
                        wait_for_calcs = True
                    try:
                        if wait_for_calcs:
                            wait_for_folder_complete(temp_tiff_folder, expected_count=20)  # Wait for the 20 files to all show up
                        # Now convert the files in the temp_tiff_folder to bitmaps
                        for file in temp_tiff_folder_path.iterdir():
                            name_key = str(file).replace('/', '\\').split("\\")[-1].split(".tif")[0]
                            arrays[name_key] = bitmap_to_array(file)
                        calculated_layers_preblend = multiple_layers_weighted(arrays)
                        calculated_layers_preblend_edge_mask = fill_border(calculated_layers_preblend, method=edge_fill)
                        calculated_layers_blended = met_ave(calculated_layers_preblend_edge_mask, radius=radius)
                        calculated_layers_blended_shrink = shrink_array(calculated_layers_blended, dat_file_9999_filled.shape)
                        calculated_layers_blended_shrink_rescale, dat_file_9999_filled_rescale, scale = rescale_to_shared_minmax(calculated_layers_blended_shrink, dat_file_9999_filled)
                        stats = align_and_compare(calculated_layers_blended_shrink_rescale, dat_file_9999_filled_rescale, ignore_zeros=False, detrend=True, with_scaling=False)
                        plot_pointclouds_and_heatmaps(calculated_layers_blended_shrink_rescale, dat_file_9999_filled_rescale, plot_save_folder, plot_save_name, stats_text=stats["text"])
                        ws.append([
                            f'{Quartile_loc}_{mat_sup_id}', Quartile_loc, edge_fill, dpi, radius, mat_sup_id, str(scale),
                            stats.get("n_points"), stats.get("rmse"), stats.get("mae"), stats.get("bias"),
                            stats.get("pearson_r"), stats.get("spearman_r"), stats.get("r2"),
                            stats.get("svd_var_explained"), stats.get("slope"), stats.get("intercept"),
                            stats.get("lambda_ratio"), stats.get("text")
                        ])
                        wb.save(excel_output_path)
                        print(f"Complete: {Quartile_loc}_{mat_sup_id}")
                    except Exception as error:
                        print(error)
                        # fig = plot_point_clouds_side_by_side_same_cmap(calculated_layers_blended_shrink_rescale, dat_file_9999_filled_rescale)
                        # fig.show()

                        pass
                    pass


print()